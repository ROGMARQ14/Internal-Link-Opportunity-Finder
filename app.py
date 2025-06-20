import streamlit as st
import pandas as pd
import numpy as np
from sklearn.metrics.pairwise import cosine_similarity
import base64
from io import BytesIO, StringIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment
import csv

def detect_csv_delimiter(file_content, sample_size=1024):
    """
    Detect the delimiter in a CSV file using multiple methods.
    Returns the detected delimiter.
    """
    # Convert bytes to string if needed
    if isinstance(file_content, bytes):
        try:
            file_content = file_content.decode('utf-8')
        except UnicodeDecodeError:
            try:
                file_content = file_content.decode('latin1')
            except:
                file_content = file_content.decode('utf-8', errors='ignore')
    
    # Method 1: Try csv.Sniffer
    try:
        sample = file_content[:sample_size]
        sniffer = csv.Sniffer()
        delimiter = sniffer.sniff(sample, delimiters=',;\t|').delimiter
        st.info(f"Detected delimiter: '{delimiter}'")
        return delimiter
    except Exception as e:
        st.warning(f"CSV Sniffer failed: {str(e)}")
    
    # Method 2: Manual detection by counting common delimiters
    sample = file_content[:sample_size]
    lines = sample.split('\n')[:5]  # Check first 5 lines
    
    delimiters = [';', ',', '\t', '|']
    delimiter_counts = {}
    
    for delimiter in delimiters:
        counts = []
        for line in lines:
            if line.strip():  # Skip empty lines
                counts.append(line.count(delimiter))
        
        # Check if delimiter appears consistently across lines
        if counts and len(set(counts)) <= 2:  # Allow some variation
            delimiter_counts[delimiter] = sum(counts) / len(counts)  # Average count
    
    if delimiter_counts:
        # Return the delimiter with highest consistent count
        best_delimiter = max(delimiter_counts, key=delimiter_counts.get)
        st.info(f"Detected delimiter: '{best_delimiter}'")
        return best_delimiter
    
    # Default fallback
    st.warning("No delimiter detected, defaulting to comma")
    return ','

def read_large_csv_safely(uploaded_file, chunk_size=10000):
    """
    Safely read large CSV files in chunks with error handling and automatic delimiter detection.
    """
    try:
        # Reset file pointer
        uploaded_file.seek(0)
        
        # Read a sample to detect delimiter
        file_content = uploaded_file.read(10240)  # Read 10KB for detection
        uploaded_file.seek(0)  # Reset again
        
        # Detect delimiter
        if isinstance(file_content, bytes):
            sample_content = file_content
        else:
            sample_content = file_content.encode()
            
        detected_delimiter = detect_csv_delimiter(sample_content)
        
        # Read in chunks
        chunks = []
        for chunk in pd.read_csv(
            uploaded_file, 
            chunksize=chunk_size,
            sep=detected_delimiter,
            on_bad_lines='skip',
            engine='python',
            encoding_errors='ignore'
        ):
            chunks.append(chunk)
        
        if not chunks:
            st.error("No data could be read from the file.")
            return None
            
        return pd.concat(chunks, ignore_index=True)
    except Exception as e:
        st.error(f"Error processing chunks: {str(e)}")
        # Try alternative methods if chunked reading fails
        try:
            uploaded_file.seek(0)
            return pd.read_csv(uploaded_file, sep=None, on_bad_lines='skip', engine='python')
        except Exception as e2:
            st.error(f"Alternative reading method failed: {str(e2)}")
            return None

def read_csv_with_auto_delimiter(uploaded_file):
    """
    Read CSV file with automatic delimiter detection and error handling.
    Tries multiple methods to ensure the file is read correctly.
    """
    try:
        # Reset file pointer
        uploaded_file.seek(0)
        file_content = uploaded_file.read()
        file_size = len(file_content)
        uploaded_file.seek(0)  # Reset again
        
        # Check file size
        st.info(f"File size: {file_size / (1024*1024):.2f} MB")
        
        # For large files, use chunked reading
        if file_size > 50 * 1024 * 1024:  # 50MB threshold
            st.info("Large file detected. Using chunked reading...")
            return read_large_csv_safely(uploaded_file)
        
        # For smaller files, try multiple methods
        
        # First, detect delimiter
        if isinstance(file_content, bytes):
            sample_content = file_content
        else:
            sample_content = file_content.encode()
            
        detected_delimiter = detect_csv_delimiter(sample_content)
        
        # Try multiple reading methods
        methods = [
            # Method 1: Use detected delimiter
            {"name": "Detected delimiter", "func": lambda: pd.read_csv(
                uploaded_file, sep=detected_delimiter, on_bad_lines='skip', engine='python')},
            # Method 2: Use pandas automatic detection
            {"name": "Pandas auto-detection", "func": lambda: pd.read_csv(
                uploaded_file, sep=None, on_bad_lines='skip', engine='python')},
            # Method 3: Try semicolon explicitly
            {"name": "Semicolon delimiter", "func": lambda: pd.read_csv(
                uploaded_file, sep=";", on_bad_lines='skip', engine='python')},
            # Method 4: Try comma explicitly
            {"name": "Comma delimiter", "func": lambda: pd.read_csv(
                uploaded_file, sep=",", on_bad_lines='skip', engine='python')},
        ]
        
        for method in methods:
            try:
                uploaded_file.seek(0)
                df = method["func"]()
                st.success(f"Successfully read CSV using {method['name']}")
                
                # Verify the data looks reasonable
                if len(df.columns) <= 1:
                    st.warning(f"Only {len(df.columns)} column detected. This may indicate an incorrect delimiter.")
                    continue
                    
                return df
            except Exception as e:
                st.warning(f"{method['name']} failed: {str(e)}")
                continue
        
        # If all methods fail, try a last resort with very permissive settings
        try:
            uploaded_file.seek(0)
            st.warning("Trying last resort parsing method...")
            df = pd.read_csv(
                uploaded_file, 
                sep=None,
                engine='python',
                on_bad_lines='skip',
                encoding_errors='ignore',
                quoting=csv.QUOTE_NONE
            )
            st.success("Successfully read CSV using last resort method")
            return df
        except Exception as e:
            st.error(f"All CSV reading methods failed. Last error: {str(e)}")
            return None
            
    except Exception as e:
        st.error(f"Error reading CSV: {str(e)}")
        return None

st.set_page_config(
    page_title="Internal Link Opportunity Finder",
    page_icon="🔗",
    layout="wide",
    initial_sidebar_state="expanded"
)

st.title("Internal Link Opportunity Finder")
st.markdown("""
This app helps you discover internal linking opportunities between related pages on your website using vector embeddings.
Based on Everett Sizemore's tutorial: [How I Found Internal Linking Opportunities With Vector Embeddings](https://moz.com/blog/internal-linking-opportunities-with-vector-embeddings)
""")

# ----------- PART 1: LINK DATASET CLEANING -----------

def clean_link_dataset(df):
    """Clean and process link dataset according to specified rules."""
    # Make a copy to avoid modifying the original
    df = df.copy()

    st.write("Initial shape:", df.shape)

    # 1. Sort by Type and filter for Hyperlinks
    if 'Type' in df.columns:
        df = df.sort_values('Type')
        df = df[df['Type'] == 'Hyperlink'].drop('Type', axis=1)
        st.write("Shape after Type filtering:", df.shape)

    # 2. Sort by Status Code and filter for 200
    if 'Status Code' in df.columns:
        df = df.sort_values('Status Code')
        df = df[df['Status Code'] == 200]
        columns_to_drop = ['Status Code', 'Status'] if 'Status' in df.columns else ['Status Code']
        df = df.drop(columns_to_drop, axis=1)
        st.write("Shape after Status filtering:", df.shape)

    # 3. Delete specified columns if they exist
    columns_to_drop = [
        'Size (Bytes)', 'Follow', 'Target', 'Rel',
        'Path Type', 'Link Path', 'Link Origin'
    ]
    columns_to_drop = [col for col in columns_to_drop if col in df.columns]

    if columns_to_drop:
        df = df.drop(columns_to_drop, axis=1)
    st.write("Remaining columns:", df.columns.tolist())

    # 4. Process Link Position if it exists
    if 'Link Position' in df.columns:
        df = df.sort_values('Link Position')
        df = df[df['Link Position'].isin(['Content', 'Aside'])]
        st.write("Shape after Link Position filtering:", df.shape)

    # 5. Clean Source URLs
    source_col = 'Source' if 'Source' in df.columns else df.columns[0]
    df = df.sort_values(source_col)

    def is_valid_page(url):
        if pd.isna(url):
            return False
        invalid_patterns = [
            'category/', 'tag/', 'sitemap', 'search', '/home/', 'index', '/page/'
        ]
        return not any(pattern in str(url).lower() for pattern in invalid_patterns)

    df = df[df[source_col].apply(is_valid_page)]
    st.write(f"Shape after {source_col} URL cleaning:", df.shape)

    # 6. Clean Destination URLs
    dest_col = 'Destination' if 'Destination' in df.columns else df.columns[1]
    df = df.sort_values(dest_col)
    df = df[df[dest_col].apply(is_valid_page)]
    st.write(f"Shape after {dest_col} URL cleaning:", df.shape)

    # 7. Process Alt Text if present
    if 'Alt Text' in df.columns and 'Anchor' in df.columns:
        df = df.sort_values('Alt Text', ascending=False)
        df.loc[df['Alt Text'].notna(), 'Anchor'] = df['Alt Text']
        df = df.drop('Alt Text', axis=1)
      
    # Clean up and standardize columns
    if 'Link Position' in df.columns:
        df = df.drop('Link Position', axis=1)

    if source_col != 'Source' or dest_col != 'Destination':
        df = df.rename(columns={source_col: 'Source', dest_col: 'Destination'})

    if 'Anchor' not in df.columns:
        df['Anchor'] = ''

    final_columns = ['Source', 'Destination', 'Anchor']
    other_columns = [col for col in df.columns if col not in final_columns]
    df = df[final_columns + other_columns]
    return df

# ----------- PART 2: EMBEDDINGS PREPROCESSING -----------

def clean_embeddings_data(df):
    """Clean and preprocess embeddings data according to specified rules."""
    st.subheader("Cleaning Embeddings Data")
    df = df.copy()
    st.write("Initial shape:", df.shape)

    # Find the embeddings column
    embeddings_col = None
    for col in df.columns:
        if 'embeddings' in col.lower() or 'extract' in col.lower():
            embeddings_col = col
            break

    if not embeddings_col:
        st.error("Could not find a column containing embeddings data. Please ensure your CSV has a column with 'embeddings' or 'extract' in its name.")
        return None

    # Sort and clean embeddings
    df = df.sort_values(embeddings_col, ascending=False)

    def is_valid_embedding(text):
        if pd.isna(text):
            return False
        invalid_words = ['timeout', 'error', 'null', 'undefined', 'nan']
        if any(word in str(text).lower() for word in invalid_words):
            return False
        text_str = str(text)
        has_numbers = any(c.isdigit() for c in text_str)
        has_separators = ',' in text_str or '.' in text_str
        return has_numbers and has_separators

    df = df[df[embeddings_col].apply(is_valid_embedding)]
    st.write("Shape after removing invalid embeddings:", df.shape)

    # Filter by status code if available
    if 'Status Code' in df.columns:
        df = df[df['Status Code'] == 200]
        st.write("Shape after status code filtering:", df.shape)

    # Determine URL column
    url_col = None
    potential_url_cols = ['URL', 'Address', 'Url', 'address']
    for col in potential_url_cols:
        if col in df.columns:
            url_col = col
            break

    if not url_col:
        for col in df.columns:
            if col != embeddings_col and col != 'Status Code' and col != 'Status':
                url_col = col
                break

    if not url_col:
        st.error("Could not identify a URL column. Please ensure your CSV has a column with URLs.")
        return None

    # Clean up and create final dataframe
    cols_to_drop = [col for col in ['Status Code', 'Status'] if col in df.columns]
    if cols_to_drop:
        df = df.drop(cols_to_drop, axis=1)

    # Add pagination filtering for URLs
    def is_valid_url(url):
        if pd.isna(url):
            return False
        invalid_patterns = [
            'category/', 'tag/', 'sitemap', 'search', '/home/', 'index', '/page/'
        ]
        return not any(pattern in str(url).lower() for pattern in invalid_patterns)

    # Apply the filtering to the URL column
    df = df[df[url_col].apply(is_valid_url)]
    st.write("Shape after filtering paginated URLs:", df.shape)

    cleaned_df = pd.DataFrame()
    cleaned_df['URL'] = df[url_col]
    cleaned_df['Embeddings'] = df[embeddings_col]

    st.write("Final embeddings data shape:", cleaned_df.shape)
    return cleaned_df

# ----------- PART 3: URL RELATIONSHIP ANALYSIS -----------

def find_related_pages(df, top_n=10):
    """Find top N related pages for each URL based on cosine similarity."""
    st.write(f"Finding top {top_n} related pages for each URL...")
    related_pages = {}
    
    with st.spinner("Processing embeddings... This may take a while for large datasets."):
        embeddings = np.stack(df['Embeddings'].values)
        urls = df['URL'].values

        # Calculate cosine similarity matrix
        cosine_similarities = cosine_similarity(embeddings)

        # For each URL, find the most similar URLs
        for idx, url in enumerate(urls):
            similar_indices = cosine_similarities[idx].argsort()[-(top_n+1):][::-1]
            similar_indices = [i for i in similar_indices if urls[i] != url][:top_n]
            related_urls = urls[similar_indices].tolist()
            related_pages[url] = related_urls

    return related_pages

def get_download_link(df, filename, link_text):
    """Create a download link for a DataFrame as Excel with formatting."""
    output = BytesIO()
    writer = pd.ExcelWriter(output, engine='openpyxl')
    df.to_excel(writer, index=False, sheet_name='URL Analysis')
    
    workbook = writer.book
    worksheet = writer.sheets['URL Analysis']
    
    # Define fills for "Exists" and "Not Found"
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    # Apply conditional formatting
    for col_idx in [4, 6, 8, 10, 12, 14, 16, 18, 20, 22]:  # Columns D, F, H, J, L in Excel (1-indexed)
        for row_idx in range(2, len(df) + 2):  # Starting from row 2 (skipping header)
            cell = worksheet.cell(row=row_idx, column=col_idx)
            if cell.value == "Exists":
                cell.fill = green_fill
            elif cell.value == "Not Found":
                cell.fill = red_fill

    # Format columns for readability
    for col in worksheet.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            cell.alignment = Alignment(vertical='center')
            if column == 'B':
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))

        if column == 'B':  # "Links to Target URL" column
            worksheet.column_dimensions[column].width = 50
        else:
            adjusted_width = max(max_length, 12)
            worksheet.column_dimensions[column].width = min(adjusted_width, 40)

    worksheet.freeze_panes = 'A2'
    writer.close()
    
    b64 = base64.b64encode(output.getvalue()).decode()
    href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{filename}">{link_text}</a>'
    return href

def get_csv_download_link(df, filename, link_text):
    """Create a download link for a DataFrame as CSV."""
    csv = df.to_csv(index=False)
    b64 = base64.b64encode(csv.encode()).decode()
    href = f'<a href="data:file/csv;base64,{b64}" download="{filename}">{link_text}</a>'
    return href

# ----------- MAIN APP STRUCTURE -----------
st.sidebar.header("About")
st.sidebar.info(
    "This app processes your internal link data and embeddings to find linking opportunities. "
    "Upload your data files, adjust settings if needed, and download the results."
)

st.sidebar.header("Instructions")
st.sidebar.markdown(
    """
    1. Export internal links from your SEO crawler (e.g., Screaming Frog)
    2. Generate vector embeddings for your pages
    3. Upload both files below
    4. Review and download the results
    """
)

st.sidebar.header("Settings")
top_n = st.sidebar.slider("Number of related URLs to find", min_value=1, max_value=10, value=10)

# Create a multi-step process with tabs
tab1, tab2, tab3 = st.tabs(["Upload Data", "Process & Analyze", "Results"])

# Global variables to store data between tabs
if 'df_cleaned_links' not in st.session_state:
    st.session_state.df_cleaned_links = None
if 'df_embeddings' not in st.session_state:
    st.session_state.df_embeddings = None
if 'final_df' not in st.session_state:
    st.session_state.final_df = None

with tab1:
    st.header("Step 1: Upload Your Data Files")
    
    st.subheader("Upload Internal Links Export")
    st.markdown("Upload the CSV file containing your internal links data (e.g., from Screaming Frog).")
    link_file = st.file_uploader("Choose a links CSV file", type=['csv'], key="link_file")
    
    if link_file is not None:
        try:
            # Use our new robust CSV reader
            df_links = read_csv_with_auto_delimiter(link_file)
            
            if df_links is not None:
                st.success(f"Successfully loaded links file with {df_links.shape[0]} rows and {df_links.shape[1]} columns")
                st.write("First few rows of the data:")
                st.dataframe(df_links.head())
                
                if st.button("Clean Links Data"):
                    with st.spinner("Cleaning links data..."):
                        st.session_state.df_cleaned_links = clean_link_dataset(df_links)
                    st.success("Links data cleaned successfully!")
                    st.dataframe(st.session_state.df_cleaned_links.head())
            else:
                st.error("Failed to load the file. Please check the file format.")
        except Exception as e:
            st.error(f"Error loading the file: {str(e)}")
    
    st.subheader("Upload Embeddings Export")
    st.markdown("Upload the CSV file containing your page embeddings.")
    embeddings_file = st.file_uploader("Choose an embeddings CSV file", type=['csv'], key="embeddings_file")
    
    if embeddings_file is not None:
        try:
            # Use our new robust CSV reader
            df_embeddings_raw = read_csv_with_auto_delimiter(embeddings_file)
            
            if df_embeddings_raw is not None:
                st.success(f"Successfully loaded embeddings file with {df_embeddings_raw.shape[0]} rows and {df_embeddings_raw.shape[1]} columns")
                st.write("First few rows of the data:")
                st.dataframe(df_embeddings_raw.head())
                
                if st.button("Clean Embeddings Data"):
                    with st.spinner("Cleaning embeddings data..."):
                        st.session_state.df_embeddings = clean_embeddings_data(df_embeddings_raw)
                    if st.session_state.df_embeddings is not None:
                        st.success("Embeddings data cleaned successfully!")
                        st.dataframe(st.session_state.df_embeddings.head())
            else:
                st.error("Failed to load the file. Please check the file format.")
        except Exception as e:
            st.error(f"Error loading the embeddings file: {str(e)}")

with tab2:
    st.header("Step 2: Process the Data")
    
    if st.session_state.df_cleaned_links is not None and st.session_state.df_embeddings is not None:
        st.success("Both datasets are loaded and cleaned. Ready to process!")
        
        if st.button("Analyze Internal Link Opportunities"):
            with st.spinner("Analyzing relationships between URLs..."):
                # Convert embeddings from string to numpy arrays
                st.session_state.df_embeddings['Embeddings'] = st.session_state.df_embeddings['Embeddings'].apply(
                    lambda x: np.array([float(i) for i in str(x).strip('[]').replace("'", "").split(',')])
                )
                
                # Find related pages
                related_pages = find_related_pages(st.session_state.df_embeddings, top_n=top_n)
                
                # Create a dictionary to store all inlinks for each URL
                inlinks_dict = {}
                for target_url in st.session_state.df_embeddings['URL']:
                    linking_sources = st.session_state.df_cleaned_links[st.session_state.df_cleaned_links['Destination'] == target_url]['Source'].tolist()
                    inlinks_dict[target_url] = ', '.join(linking_sources) if linking_sources else ''
                
                # Create the final DataFrame
                output_data = []
                
                for url, related_urls in related_pages.items():
                    padded_related_urls = related_urls + [None] * (top_n - len(related_urls))
                    links_to_target = inlinks_dict.get(url, '')
                    
                    if not links_to_target:
                        links_to_target = "none"
                    
                    row = {
                        'Target URL': url,
                        'Links to Target URL': links_to_target
                    }
                    
                    for i, related_url in enumerate(padded_related_urls, 1):
                        row[f'Related URL {i}'] = related_url
                        if related_url is not None:
                            exists_status = "Exists" if related_url in st.session_state.df_cleaned_links[st.session_state.df_cleaned_links['Destination'] == url]['Source'].values else "Not Found"
                            row[f'URL {i} links to A?'] = exists_status
                        else:
                            row[f'URL {i} links to A?'] = "Not Found"
                    
                    output_data.append(row)
                
                # Create the final DataFrame with columns in the correct order
                column_order = ['Target URL', 'Links to Target URL']
                for i in range(1, top_n + 1):
                    column_order.extend([f'Related URL {i}', f'URL {i} links to A?'])
                
                st.session_state.final_df = pd.DataFrame(output_data)
                st.session_state.final_df = st.session_state.final_df[column_order]
            
            st.success("Analysis completed successfully!")
            st.subheader("Analysis Summary")
            st.write(f"Found {len(st.session_state.final_df)} target URLs with related pages")
            st.write(f"Identified {st.session_state.final_df['URL 1 links to A?'].value_counts().get('Not Found', 0)} potential primary linking opportunities")
    else:
        st.warning("Please complete Step 1 first. Upload and clean both datasets before proceeding.")

with tab3:
    st.header("Step 3: Review and Download Results")
    
    if st.session_state.final_df is not None:
        st.subheader("Internal Link Opportunities")
        
        # Add filters for the results
        st.markdown("**Filter Results**")
        col1, col2 = st.columns(2)
        with col1:
            filter_exists = st.checkbox("Show only pages with missing links", value=True)
        with col2:
            search_term = st.text_input("Filter by URL containing:")
        
        # Apply filters
        filtered_df = st.session_state.final_df.copy()
        if filter_exists:
            mask = False
            for i in range(1, top_n + 1):
                mask = mask | (filtered_df[f'URL {i} links to A?'] == 'Not Found')
            filtered_df = filtered_df[mask]
        
        if search_term:
            mask = filtered_df['Target URL'].str.contains(search_term, case=False, na=False)
            for i in range(1, top_n + 1):
                mask = mask | filtered_df[f'Related URL {i}'].str.contains(search_term, case=False, na=False)
            filtered_df = filtered_df[mask]
        
        # Display the filtered results
        st.dataframe(filtered_df, use_container_width=True)
        
        st.subheader("Download Results")
        col1, col2 = st.columns(2)
        with col1:
            st.markdown(get_download_link(filtered_df, 'internal_link_opportunities.xlsx', 'Download Excel File'), unsafe_allow_html=True)
        with col2:
            st.markdown(get_csv_download_link(filtered_df, 'internal_link_opportunities.csv', 'Download CSV File'), unsafe_allow_html=True)
        
        st.subheader("Opportunity Analysis")
        
        # Calculate and display stats about the opportunities
        total_opportunities = sum(
            (filtered_df[f'URL {i} links to A?'] == 'Not Found').sum() 
            for i in range(1, top_n + 1)
        )
        total_pages = len(filtered_df)
        
        st.write(f"Found {total_opportunities} total linking opportunities across {total_pages} pages")
        
        # Show a chart of opportunities by page
        opportunities_by_page = []
        for _, row in filtered_df.iterrows():
            count = sum(1 for i in range(1, top_n + 1) if row[f'URL {i} links to A?'] == 'Not Found')
            opportunities_by_page.append(count)
        
        opportunity_counts = pd.Series(opportunities_by_page).value_counts().sort_index()
        
        st.write("Distribution of opportunities per page:")
        st.bar_chart(opportunity_counts)
        
        # Show pages with most opportunities
        st.subheader("Top Pages to Add Links To")
        top_pages = filtered_df.copy()
        top_pages['Missing Links'] = [sum(1 for i in range(1, top_n + 1) if row[f'URL {i} links to A?'] == 'Not Found') for _, row in top_pages.iterrows()]
        top_pages = top_pages.sort_values('Missing Links', ascending=False)[['Target URL', 'Missing Links']].head(10)
        st.table(top_pages)
    else:
        st.warning("No results yet. Please complete Steps 1 and 2 first.")

# Footer
st.markdown("---")
st.markdown(
    """
    <div style="text-align: center">
        <p style="color: #888">
            Based on Everett Sizemore's tutorial: 
            <a href="https://moz.com/blog/internal-linking-opportunities-with-vector-embeddings" target="_blank">
                How I Found Internal Linking Opportunities With Vector Embeddings
            </a>
        </p>
    </div>
    """, 
    unsafe_allow_html=True
)
