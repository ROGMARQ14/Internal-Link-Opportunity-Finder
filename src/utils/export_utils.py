"""
Export utilities for generating Excel and CSV files with formatting.
"""

import pandas as pd
import base64
import logging
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Alignment, Font
from openpyxl.utils.dataframe import dataframe_to_rows
from typing import Dict, Any, List, Optional

logger = logging.getLogger(__name__)

class ExportFormatter:
    """Handles formatting and export of analysis results."""

    def __init__(self):
        # Define color schemes
        self.colors = {
            'exists': PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"),
            'not_found': PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"),
            'high_quality': PatternFill(start_color="D4E6F1", end_color="D4E6F1", fill_type="solid"),
            'low_quality': PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid"),
            'header': PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
        }

        self.fonts = {
            'header': Font(color="FFFFFF", bold=True),
            'bold': Font(bold=True)
        }

    def create_excel_with_formatting(self, df: pd.DataFrame, filename: str) -> str:
        """
        Create a formatted Excel file with conditional formatting.

        Args:
            df: DataFrame to export
            filename: Output filename

        Returns:
            Base64 encoded Excel file
        """
        try:
            output = BytesIO()

            # Create workbook and worksheet
            wb = Workbook()
            ws = wb.active
            ws.title = "Link Analysis Results"

            # Add data to worksheet
            for r in dataframe_to_rows(df, index=False, header=True):
                ws.append(r)

            # Apply header formatting
            header_row = 1
            for col_num, cell in enumerate(ws[header_row], 1):
                cell.fill = self.colors['header']
                cell.font = self.fonts['header']
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Apply conditional formatting for link status columns
            link_status_columns = self._find_link_status_columns(df)
            for col_idx in link_status_columns:
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value == "Exists":
                        cell.fill = self.colors['exists']
                    elif cell.value == "Not Found":
                        cell.fill = self.colors['not_found']

            # Apply quality score formatting
            quality_columns = self._find_quality_columns(df)
            for col_idx in quality_columns:
                for row_idx in range(2, len(df) + 2):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if cell.value is not None:
                        try:
                            score = float(cell.value)
                            if score > 0.7:
                                cell.fill = self.colors['high_quality']
                            elif score < 0.3:
                                cell.fill = self.colors['low_quality']
                        except (ValueError, TypeError):
                            continue

            # Auto-adjust column widths
            self._adjust_column_widths(ws, df)

            # Freeze top row
            ws.freeze_panes = 'A2'

            # Save workbook
            wb.save(output)

            # Encode as base64
            b64_data = base64.b64encode(output.getvalue()).decode()

            logger.info(f"Created formatted Excel file: {filename}")
            return b64_data

        except Exception as e:
            logger.error(f"Error creating Excel file: {e}")
            raise

    def _find_link_status_columns(self, df: pd.DataFrame) -> List[int]:
        """Find columns that contain link status information."""
        status_columns = []
        for col_idx, col_name in enumerate(df.columns, 1):
            if "links to A?" in col_name:
                status_columns.append(col_idx)
        return status_columns

    def _find_quality_columns(self, df: pd.DataFrame) -> List[int]:
        """Find columns that contain quality scores."""
        quality_columns = []
        for col_idx, col_name in enumerate(df.columns, 1):
            if any(keyword in col_name.lower() for keyword in ['quality', 'score', 'opportunity']):
                quality_columns.append(col_idx)
        return quality_columns

    def _adjust_column_widths(self, ws, df: pd.DataFrame):
        """Auto-adjust column widths based on content."""
        for col_idx, col_name in enumerate(df.columns, 1):
            column_letter = ws.cell(row=1, column=col_idx).column_letter

            # Calculate max width
            max_length = len(str(col_name))
            for row_idx in range(2, len(df) + 2):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))

            # Set column width with limits
            if "URL" in col_name:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 60)
            elif "Links to Target URL" in col_name:
                ws.column_dimensions[column_letter].width = 50
            else:
                ws.column_dimensions[column_letter].width = min(max_length + 2, 30)

    def create_summary_sheet(self, wb: Workbook, df: pd.DataFrame, 
                           link_stats: Dict[str, Any]) -> None:
        """
        Create a summary sheet with key metrics.

        Args:
            wb: Workbook object
            df: Analysis results DataFrame
            link_stats: Link statistics
        """
        # Create summary worksheet
        summary_ws = wb.create_sheet("Summary", 0)

        # Add title
        summary_ws['A1'] = "Internal Link Analysis Summary"
        summary_ws['A1'].font = Font(size=16, bold=True)
        summary_ws.merge_cells('A1:D1')

        # Add metrics
        metrics = [
            ("Total Target URLs", len(df)),
            ("Total Internal Links", link_stats.get('total_links', 0)),
            ("Unique Source Pages", link_stats.get('unique_sources', 0)),
            ("Unique Destination Pages", link_stats.get('unique_destinations', 0)),
            ("Average Outlinks per Source", f"{link_stats.get('avg_outlinks_per_source', 0):.1f}"),
            ("Average Inlinks per Destination", f"{link_stats.get('avg_inlinks_per_destination', 0):.1f}"),
        ]

        for row_idx, (metric, value) in enumerate(metrics, 3):
            summary_ws[f'A{row_idx}'] = metric
            summary_ws[f'B{row_idx}'] = value
            summary_ws[f'A{row_idx}'].font = self.fonts['bold']

        # Calculate opportunity metrics
        opportunity_metrics = self._calculate_opportunity_metrics(df)

        # Add opportunity analysis
        summary_ws['A12'] = "Opportunity Analysis"
        summary_ws['A12'].font = Font(size=14, bold=True)

        for row_idx, (metric, value) in enumerate(opportunity_metrics, 14):
            summary_ws[f'A{row_idx}'] = metric
            summary_ws[f'B{row_idx}'] = value
            summary_ws[f'A{row_idx}'].font = self.fonts['bold']

        # Auto-adjust column widths
        summary_ws.column_dimensions['A'].width = 35
        summary_ws.column_dimensions['B'].width = 15

    def _calculate_opportunity_metrics(self, df: pd.DataFrame) -> List[tuple]:
        """Calculate opportunity metrics from the analysis results."""
        metrics = []

        # Count missing links
        missing_links = 0
        total_opportunities = 0

        for col in df.columns:
            if "links to A?" in col:
                missing_links += (df[col] == "Not Found").sum()
                total_opportunities += df[col].notna().sum()

        metrics.append(("Total Link Opportunities", total_opportunities))
        metrics.append(("Missing Links Found", missing_links))

        if total_opportunities > 0:
            miss_rate = (missing_links / total_opportunities) * 100
            metrics.append(("Missing Link Rate", f"{miss_rate:.1f}%"))

        # Calculate average quality scores if available
        quality_cols = [col for col in df.columns if "Quality Score" in col]
        if quality_cols:
            avg_quality = df[quality_cols].mean().mean()
            metrics.append(("Average Link Quality", f"{avg_quality:.2f}"))

        return metrics

class ExportManager:
    """Manages export functionality for the application."""

    def __init__(self):
        self.formatter = ExportFormatter()

    def get_excel_download_link(self, df: pd.DataFrame, filename: str, 
                               link_text: str, link_stats: Optional[Dict[str, Any]] = None) -> str:
        """
        Create a download link for a DataFrame as formatted Excel.

        Args:
            df: DataFrame to export
            filename: Output filename
            link_text: Text for the download link
            link_stats: Optional link statistics for summary sheet

        Returns:
            HTML download link
        """
        try:
            b64_data = self.formatter.create_excel_with_formatting(df, filename)

            href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64_data}" download="{filename}">{link_text}</a>'
            return href

        except Exception as e:
            logger.error(f"Error creating Excel download link: {e}")
            return f"<span style='color: red;'>Error creating Excel file</span>"

    def get_csv_download_link(self, df: pd.DataFrame, filename: str, link_text: str) -> str:
        """
        Create a download link for a DataFrame as CSV.

        Args:
            df: DataFrame to export
            filename: Output filename
            link_text: Text for the download link

        Returns:
            HTML download link
        """
        try:
            csv_data = df.to_csv(index=False)
            b64_data = base64.b64encode(csv_data.encode()).decode()

            href = f'<a href="data:text/csv;base64,{b64_data}" download="{filename}">{link_text}</a>'
            return href

        except Exception as e:
            logger.error(f"Error creating CSV download link: {e}")
            return f"<span style='color: red;'>Error creating CSV file</span>"

    def create_detailed_report(self, df: pd.DataFrame, link_stats: Dict[str, Any], 
                             filename: str = "detailed_link_analysis.xlsx") -> str:
        """
        Create a detailed report with multiple sheets.

        Args:
            df: Analysis results DataFrame
            link_stats: Link statistics
            filename: Output filename

        Returns:
            Base64 encoded Excel file
        """
        try:
            output = BytesIO()

            # Create workbook
            wb = Workbook()

            # Remove default sheet
            wb.remove(wb.active)

            # Create summary sheet
            self.formatter.create_summary_sheet(wb, df, link_stats)

            # Create main analysis sheet
            main_ws = wb.create_sheet("Detailed Analysis")
            for r in dataframe_to_rows(df, index=False, header=True):
                main_ws.append(r)

            # Apply formatting to main sheet
            self.formatter._adjust_column_widths(main_ws, df)

            # Create opportunities-only sheet
            opportunities_df = self._filter_opportunities(df)
            if len(opportunities_df) > 0:
                opp_ws = wb.create_sheet("Opportunities Only")
                for r in dataframe_to_rows(opportunities_df, index=False, header=True):
                    opp_ws.append(r)
                self.formatter._adjust_column_widths(opp_ws, opportunities_df)

            # Save workbook
            wb.save(output)

            # Encode as base64
            b64_data = base64.b64encode(output.getvalue()).decode()

            logger.info(f"Created detailed report: {filename}")
            return b64_data

        except Exception as e:
            logger.error(f"Error creating detailed report: {e}")
            raise

    def _filter_opportunities(self, df: pd.DataFrame) -> pd.DataFrame:
        """Filter DataFrame to show only rows with opportunities."""
        mask = pd.Series([False] * len(df))

        for col in df.columns:
            if "links to A?" in col:
                mask = mask | (df[col] == "Not Found")

        return df[mask]
